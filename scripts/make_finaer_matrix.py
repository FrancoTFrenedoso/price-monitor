from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.cell.cell import MergedCell


PLAZOS = [3, 6, 12, 24, 36]


def segmento(alq_exp: float) -> str:
    if alq_exp <= 500_000:
        return "hasta 500k"
    if alq_exp <= 800_000:
        return "500k-800k"
    return "mayor_800k"


def pick_plan(planes: list[dict], mode: str) -> dict:
    if not planes:
        return {}
    if mode == "contado":
        p1 = [
            p for p in planes
            if int(p.get("cuotas") or p.get("cantidad_de_cuotas") or 0) == 1
        ]
        if p1:
            return p1[0]
    return sorted(planes, key=lambda p: float(p.get("monto_final") or 1e18))[0]


def load_latest_jsonl() -> Path:
    out = Path("output")
    files = sorted(out.glob("finaer_*.jsonl"))
    if not files:
        raise SystemExit("No hay output/finaer_*.jsonl (corré primero: python -m price_monitor.cli)")
    return files[-1]


def to_float(x) -> Optional[float]:
    try:
        if x is None or pd.isna(x):
            return None
        return float(x)
    except Exception:
        return None


def main(mode: str = "contado"):
    jsonl_path = load_latest_jsonl()
    rows = []

    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rec = json.loads(line)

            s = rec["scenario"]
            alq = float(s["alquiler"])
            exp = float(s.get("expensas") or 0)
            meses = int(s["meses"])
            alq_exp = alq + exp

            planes = rec["normalized"]["planes"]
            p = pick_plan(planes, mode=mode)
            if not p:
                continue

            monto_final = to_float(p.get("monto_final"))
            if monto_final is None or alq_exp <= 0:
                continue

            # fracción: monto_final / (alq+exp)
            pct_ae = monto_final / alq_exp

            # porcentaje (0-100+)
            pct_ae_pct = pct_ae * 100.0

            # descuento real si viene (lo normalizamos a %)
            pct_desc = to_float(p.get("pct_descuento_real"))
            if pct_desc is not None and pct_desc <= 1:
                pct_desc = pct_desc * 100.0

            rows.append({
                "segmento": segmento(alq_exp),
                "meses": meses,
                "alq_exp": alq_exp,
                "monto_final": monto_final,
                "pct_ae_pct": pct_ae_pct,     # para matriz
                "pct_desc_pct": pct_desc,     # para hoja descuentos
            })

    df = pd.DataFrame(rows)
    df = df[df["meses"].isin(PLAZOS)].copy()

    if df.empty:
        raise SystemExit("No hay filas válidas para construir matriz (revisá escenarios/plazos y normalized).")

    g = df.groupby(["segmento", "meses"], as_index=False).agg(
        pct_ae_prom=("pct_ae_pct", "mean"),
        pct_desc_prom=("pct_desc_pct", "mean"),
        n=("monto_final", "count"),
    )

    mat = (
        g.pivot(index="segmento", columns="meses", values="pct_ae_prom")
        .reindex(columns=PLAZOS)
        .reset_index()
    )

    des = g.groupby("meses", as_index=False).agg(desc_prom=("pct_desc_prom", "mean"))
    des_row: dict[str, float | str | None] = {"segmento": "Des."}
    for m in PLAZOS:
        v = des.loc[des["meses"] == m, "desc_prom"]
        des_row[str(m)] = float(v.iloc[0]) if len(v) and pd.notna(v.iloc[0]) else None
    des_df = pd.DataFrame([des_row])

    out_xlsx = jsonl_path.with_name(f"finaer_matrix_pct_{jsonl_path.stem}_{mode}.xlsx")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        mat.to_excel(w, index=False, sheet_name="Matriz_Finaer_%AE")
        des_df.to_excel(w, index=False, sheet_name="Descuentos_Finaer")
        df.to_excel(w, index=False, sheet_name="Base")

    # formato Excel: encabezado + porcentajes sin romper merged cells
    wb = load_workbook(out_xlsx)
    ws = wb["Matriz_Finaer_%AE"]

    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Convertir valores (ej 60) -> fracción (0.60) y aplicar formato %
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)

            # saltar celdas merged (read-only)
            if isinstance(cell, MergedCell) or cell.coordinate in ws.merged_cells:
                continue

            val = cell.value
            if isinstance(val, (int, float)):
                cell.value = float(val) / 100.0
                cell.number_format = "0.00%"

    # color scale
    max_row = ws.max_row
    max_col = ws.max_column
    if max_col >= 2 and max_row >= 2:
        from openpyxl.utils import get_column_letter
        rng = f"B2:{get_column_letter(max_col)}{max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            )
        )

    wb.save(out_xlsx)
    print("Wrote", out_xlsx)


if __name__ == "__main__":
    main(mode="contado")  # o "min_total"
