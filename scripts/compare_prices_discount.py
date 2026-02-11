# scripts/compare_prices_discount.py
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional, cast

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet



# ---------------- Config ----------------
PLAZOS_ALL = [3, 6, 12, 24, 36]
PLAZOS_FINAER = [12, 24, 36]  # tu crawler de Finaer hoy corre estos

# regla transferencia: 15% OFF SOLO para 1 pago
TRANSFER_DESC_PCT = 15.0

# CSV "por web" de Hoggax (24/36). Ruta real en tu repo:
HOGGAX_WEB_CSV_LONG = Path("src/price_monitor/data/hoggax_rates_long.csv")


# ---------------- Helpers ----------------
def to_float(x) -> Optional[float]:
    try:
        if x is None or pd.isna(x):
            return None
        return float(x)
    except Exception:
        return None


def segmento(alq_exp: float) -> str:
    if alq_exp <= 500_000:
        return "hasta 500k"
    if alq_exp <= 800_000:
        return "500k-800k"
    return "mayor_800k"


def resolve_repo_relative(p: str | Path) -> Path:
    # scripts/compare_prices_discount.py -> repo_root = parents[1]
    base_dir = Path(__file__).resolve().parents[1]
    pp = Path(p)
    return pp if pp.is_absolute() else (base_dir / pp)


def load_latest_jsonl() -> Path:
    files = sorted(Path("output").glob("finaer_*.jsonl"))
    if not files:
        raise SystemExit("No hay output/finaer_*.jsonl. Corré primero: python -m price_monitor.cli")
    return files[-1]


def as_int(x, default: int = 0) -> int:
    try:
        if x is None or pd.isna(x):
            return default
        return int(float(x))
    except Exception:
        return default


def as_money(x) -> Optional[float]:
    try:
        if x is None or pd.isna(x):
            return None
        return float(x)
    except Exception:
        return None


def load_hoggax_web_long(path: str | Path = HOGGAX_WEB_CSV_LONG) -> pd.DataFrame:
    """
    CSV long esperado:
      segmento,plazo_meses,precio_garantia,porcentaje_sobre_total,descuento_pct

    Interpretación en este comparador:
    - precio_garantia = PRECIO LISTA por web para 24/36 (antes de transferencia)
    """
    p = resolve_repo_relative(path)
    if not p.exists():
        raise SystemExit(f"No existe {p}. Necesitás el CSV long de Hoggax para 24/36.")
    df = pd.read_csv(p)

    df["segmento"] = df["segmento"].astype(str).str.strip()
    df["plazo_meses"] = pd.to_numeric(df["plazo_meses"], errors="coerce").astype("Int64")
    df["precio_garantia"] = pd.to_numeric(df["precio_garantia"], errors="coerce")

    df = df.rename(columns={"precio_garantia": "hoggax_precio_lista_web"})
    return df[["segmento", "plazo_meses", "hoggax_precio_lista_web"]]


# ---------------- Excel formatting ----------------
def style_header_row(ws, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = cast(Cell, ws.cell(row=row, column=col))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def heatmap(ws, rng: str):
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min",
            start_color="63BE7B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="F8696B",
        ),
    )


def money(cell: Cell):
    cell.number_format = '"$"#,##0'
    cell.alignment = Alignment(horizontal="center")


def pct(cell: Cell):
    cell.number_format = "0.00%"
    cell.alignment = Alignment(horizontal="center")


# ---------------- Main ----------------
def main():
    jsonl_path = load_latest_jsonl()

    # 1) FINaer: construir tabla por MISMA ENTRADA (alquiler, expensas, meses) y por PLAN (cuotas)
    finaer_rows: list[dict] = []
    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rec = json.loads(line)
            s = rec.get("scenario") or {}

            alquiler = float(s.get("alquiler") or 0)
            expensas = float(s.get("expensas") or 0)
            meses = int(s.get("meses") or 0)
            alq_exp = alquiler + expensas

            if alq_exp <= 0 or meses not in PLAZOS_FINAER:
                continue

            seg = segmento(alq_exp)
            planes = (rec.get("normalized") or {}).get("planes") or []
            if not isinstance(planes, list) or not planes:
                continue

            for p in planes:
                cuotas = as_int(p.get("cantidad_de_cuotas", p.get("cuotas")), default=0)
                if cuotas <= 0:
                    continue

                lista = as_money(p.get("honorario_sin_descuentos"))
                if lista is None or lista <= 0:
                    continue

                finaer_total_web = as_money(p.get("monto_final"))
                finaer_anticipo_web = as_money(p.get("anticipo"))
                finaer_monto_cuota_web = as_money(p.get("monto_cuotas"))

                # regla transferencia 15% SOLO para 1 pago:
                finaer_transfer_desc_pct = TRANSFER_DESC_PCT if cuotas == 1 else 0.0
                finaer_total_transfer = lista * (1.0 - finaer_transfer_desc_pct / 100.0) if cuotas == 1 else finaer_total_web

                finaer_rows.append(
                    dict(
                        alquiler=alquiler,
                        expensas=expensas,
                        alq_exp=alq_exp,
                        segmento=seg,
                        plazo_meses=meses,
                        cuotas=cuotas,
                        finaer_precio_lista=lista,
                        finaer_total_web=finaer_total_web,
                        finaer_anticipo_web=finaer_anticipo_web,
                        finaer_monto_cuota_web=finaer_monto_cuota_web,
                        finaer_transfer_desc_pct=finaer_transfer_desc_pct,
                        finaer_total_transfer=finaer_total_transfer,
                    )
                )

    df_f = pd.DataFrame(finaer_rows)
    if df_f.empty:
        raise SystemExit("Finaer: no hay datos. Corré la CLI con escenarios 12/24/36 y planes.")

    # Si corriste múltiples escenarios iguales, promediá por misma entrada + plan
    key = ["alquiler", "expensas", "alq_exp", "segmento", "plazo_meses", "cuotas"]
    df_f = (
        df_f.groupby(key, as_index=False)
        .agg(
            finaer_precio_lista=("finaer_precio_lista", "mean"),
            finaer_total_web=("finaer_total_web", "mean"),
            finaer_anticipo_web=("finaer_anticipo_web", "mean"),
            finaer_monto_cuota_web=("finaer_monto_cuota_web", "mean"),
            finaer_transfer_desc_pct=("finaer_transfer_desc_pct", "mean"),
            finaer_total_transfer=("finaer_total_transfer", "mean"),
        )
        .sort_values(by=["segmento", "plazo_meses", "alq_exp", "cuotas"])
        .reset_index(drop=True)
    )

    # 2) Hoggax: construir desde la MISMA ENTRADA
    # 24/36 por web CSV (por segmento/plazo; si querés exacto por input, el CSV debe incluir alq_exp)
    df_h_web = load_hoggax_web_long(HOGGAX_WEB_CSV_LONG)

    df = df_f.copy()

    alq = pd.to_numeric(df["alq_exp"], errors="coerce").astype(float)
    plazo = pd.to_numeric(df["plazo_meses"], errors="coerce").astype("Int64")

    # reglas Pablo (lista) para 3/6/12
    hoggax_lista_regla = np.where(
        plazo.eq(12),
        alq / 0.9,
        np.where(plazo.isin([3, 6]), alq * 0.8, np.nan),
    )
    df["hoggax_precio_lista_regla"] = hoggax_lista_regla

    # merge web (24/36)
    df = df.merge(df_h_web, on=["segmento", "plazo_meses"], how="left")

    df["hoggax_precio_lista"] = df["hoggax_precio_lista_regla"]
    df.loc[df["hoggax_precio_lista"].isna(), "hoggax_precio_lista"] = df["hoggax_precio_lista_web"]

    # transferencia 15% solo 1 pago
    cuotas = pd.to_numeric(df["cuotas"], errors="coerce").astype("Int64")
    df["hoggax_transfer_desc_pct"] = np.where(cuotas.eq(1), TRANSFER_DESC_PCT, 0.0)
    df["hoggax_total_transfer"] = df["hoggax_precio_lista"] * (1.0 - (df["hoggax_transfer_desc_pct"] / 100.0))

    # desglose simple (si no tenés financiamiento real de Hoggax en cuotas)
    df["hoggax_monto_cuota_teorico"] = np.where(
        cuotas.gt(1) & pd.notna(df["hoggax_total_transfer"]),
        df["hoggax_total_transfer"] / cuotas.astype(float),
        0.0,
    )
    df["hoggax_anticipo_teorico"] = np.where(
        cuotas.gt(0) & pd.notna(df["hoggax_total_transfer"]),
        df["hoggax_total_transfer"] / cuotas.astype(float),
        np.nan,
    )

    # 3) Diferencias (comparables)
    df["dif_lista_$"] = df["finaer_precio_lista"] - df["hoggax_precio_lista"]
    df["dif_total_transfer_$"] = df["finaer_total_transfer"] - df["hoggax_total_transfer"]

    # 4) Excel
    out = Path("output") / f"compare_same_input_{jsonl_path.stem}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = cast(Worksheet, wb.active)

    ws.title = "Comparativa"

    headers = [
        "alquiler",
        "expensas",
        "alq_exp",
        "segmento",
        "plazo_meses",
        "cuotas",
        "finaer_precio_lista",
        "finaer_total_web",
        "finaer_transfer_desc_pct",
        "finaer_total_transfer",
        "finaer_anticipo_web",
        "finaer_monto_cuota_web",
        "hoggax_precio_lista",
        "hoggax_transfer_desc_pct",
        "hoggax_total_transfer",
        "hoggax_anticipo_teorico",
        "hoggax_monto_cuota_teorico",
        "dif_lista_$",
        "dif_total_transfer_$",
    ]
    ws.append(headers)
    style_header_row(ws, 1, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 11,
        "F": 8,
        "G": 18,
        "H": 16,
        "I": 20,
        "J": 18,
        "K": 16,
        "L": 18,
        "M": 18,
        "N": 20,
        "O": 18,
        "P": 18,
        "Q": 20,
        "R": 14,
        "S": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for _, r in df.iterrows():
        ws.append(
            [
                to_float(r["alquiler"]),
                to_float(r["expensas"]),
                to_float(r["alq_exp"]),
                r["segmento"],
                int(r["plazo_meses"]),
                int(r["cuotas"]),
                to_float(r["finaer_precio_lista"]),
                to_float(r["finaer_total_web"]),
                to_float(r["finaer_transfer_desc_pct"]),
                to_float(r["finaer_total_transfer"]),
                to_float(r["finaer_anticipo_web"]),
                to_float(r["finaer_monto_cuota_web"]),
                to_float(r["hoggax_precio_lista"]),
                to_float(r["hoggax_transfer_desc_pct"]),
                to_float(r["hoggax_total_transfer"]),
                to_float(r["hoggax_anticipo_teorico"]),
                to_float(r["hoggax_monto_cuota_teorico"]),
                to_float(r["dif_lista_$"]),
                to_float(r["dif_total_transfer_$"]),
            ]
        )

    for row in range(2, ws.max_row + 1):
        # dinero
        for col in (1, 2, 3, 7, 8, 10, 11, 12, 13, 15, 16, 17, 18, 19):
            cell = cast(Cell, ws.cell(row, col))
            if isinstance(cell, MergedCell) or cell.coordinate in ws.merged_cells:
                continue
            money(cell)

        # % (transfer)
        for col in (9, 14):
            cell = cast(Cell, ws.cell(row, col))
            if isinstance(cell, MergedCell) or cell.coordinate in ws.merged_cells:
                continue
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.value = float(v) / 100.0
            pct(cell)

        ws.cell(row, 5).alignment = Alignment(horizontal="center")
        ws.cell(row, 6).alignment = Alignment(horizontal="center")

    heatmap(ws, f"S2:S{ws.max_row}")  # dif_total_transfer_$

    wb.save(out)
    print("Wrote Excel ->", out)


if __name__ == "__main__":
    main()
