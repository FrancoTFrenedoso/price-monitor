# scripts/make_summary_compare.py
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional, cast

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Mapping



# Hoggax: 3/6/12/24/36 (según tu CSV manual)
PLAZOS_HOGGAX = [3, 6, 12, 24, 36]
# Finaer: NO existe 3/6 (según tu aclaración)
PLAZOS_FINAER = [12, 24, 36]

SEGMENTOS = ["hasta 500k", "500k-800k", "mayor_800k"]

SENTINEL_CUOTAS = 10**9


def to_float(x) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x):
            return None
        return float(x)
    except Exception:
        return None


def segmento(alq_exp: float) -> str:
    if alq_exp <= 500_000:
        return "hasta 500k"
    if alq_exp <= 800_000:
        return "500k-800k"
    return "mayor_800k"


def load_latest_jsonl(prefix: str = "finaer_") -> Path:
    files = sorted(Path("output").glob(f"{prefix}*.jsonl"))
    if not files:
        raise SystemExit(f"No hay output/{prefix}*.jsonl. Corré primero: python -m price_monitor.cli")
    return files[-1]


def cuotas_of(p: Mapping[str, Any]) -> int:
    c = p.get("cuotas")
    if c is None:
        c = p.get("cantidad_de_cuotas")

    if c is None:
        return SENTINEL_CUOTAS

    # a veces viene "12", 12.0, "12 cuotas", etc.
    if isinstance(c, bool):
        return SENTINEL_CUOTAS

    try:
        if isinstance(c, str):
            c = c.strip()
            # opcional: quedarte con los dígitos iniciales
            # ejemplo "12 cuotas" -> "12"
            num = ""
            for ch in c:
                if ch.isdigit():
                    num += ch
                elif num:
                    break
            return int(num) if num else SENTINEL_CUOTAS

        return int(c)
    except Exception:
        return SENTINEL_CUOTAS


def pick_plan_contado(planes: list[Mapping[str, Any]]) -> Mapping[str, Any]:
    if not planes:
        return {}

    # 1) cuota=1 gana
    ones = [p for p in planes if cuotas_of(p) == 1]
    if ones:
        return ones[0]

    # 2) si no hay 1 cuota: menor monto_final
    def monto_key(p: Mapping[str, Any]) -> float:
        v = p.get("monto_final")
        try:
            return float(v) if v is not None else 1e18
        except Exception:
            return 1e18

    return sorted(planes, key=monto_key)[0]


def load_hoggax_rates_long(path: str) -> pd.DataFrame:
    """
    Soporta dos formatos:

    1) LONG (completo):
       segmento,plazo_meses,hoggax_precio_$,hoggax_desc_pct,hoggax_pct_sobre_total,hoggax_cuota_$,hoggax_cuotas

    2) WIDE (solo % sobre total):
       segmento,3,6,12,24,36
       (donde cada columna de plazo contiene hoggax_pct_sobre_total)
    """
    df = pd.read_csv(path)

    # normalizar nombres
    df = df.rename(columns={c: str(c).strip() for c in df.columns})

    # ---- caso WIDE: columnas '3','6','12','24','36' ----
    wide_plazos = [c for c in df.columns if str(c).strip().isdigit()]
    if "segmento" in df.columns and wide_plazos:
        df["segmento"] = df["segmento"].astype(str).str.strip()

        # melt a long
        df_long = df.melt(
            id_vars=["segmento"],
            value_vars=wide_plazos,
            var_name="plazo_meses",
            value_name="hoggax_pct_sobre_total",
        )

        df_long["plazo_meses"] = pd.to_numeric(df_long["plazo_meses"], errors="coerce").astype("Int64")
        df_long["hoggax_pct_sobre_total"] = pd.to_numeric(df_long["hoggax_pct_sobre_total"], errors="coerce")

        # normalizar a %
        vals = df_long["hoggax_pct_sobre_total"].dropna()
        if len(vals) and float(vals.max()) <= 3.5:
            df_long["hoggax_pct_sobre_total"] = df_long["hoggax_pct_sobre_total"] * 100.0

        # completar columnas faltantes (no existen en wide)
        df_long["hoggax_precio_$"] = pd.NA
        df_long["hoggax_desc_pct"] = pd.NA
        df_long["hoggax_cuota_$"] = pd.NA
        df_long["hoggax_cuotas"] = pd.NA

        return df_long[
            [
                "segmento",
                "plazo_meses",
                "hoggax_precio_$",
                "hoggax_desc_pct",
                "hoggax_pct_sobre_total",
                "hoggax_cuota_$",
                "hoggax_cuotas",
            ]
        ]

    # ---- caso LONG: validar columnas requeridas ----
    required = [
        "segmento",
        "plazo_meses",
        "hoggax_precio_$",
        "hoggax_desc_pct",
        "hoggax_pct_sobre_total",
        "hoggax_cuota_$",
        "hoggax_cuotas",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"CSV Hoggax inválido. Faltan columnas: {missing}. Columnas presentes: {list(df.columns)}"
        )

    df["segmento"] = df["segmento"].astype(str).str.strip()
    df["plazo_meses"] = pd.to_numeric(df["plazo_meses"], errors="coerce").astype("Int64")

    for c in ["hoggax_precio_$", "hoggax_desc_pct", "hoggax_pct_sobre_total", "hoggax_cuota_$", "hoggax_cuotas"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # desc_pct: puede venir 0.15 o 15 -> %
    vals = df["hoggax_desc_pct"].dropna()
    if len(vals) and float(vals.max()) <= 3.5:
        df["hoggax_desc_pct"] = df["hoggax_desc_pct"] * 100.0

    # pct_sobre_total: puede venir 0.06 o 6 -> %
    vals2 = df["hoggax_pct_sobre_total"].dropna()
    if len(vals2) and float(vals2.max()) <= 3.5:
        df["hoggax_pct_sobre_total"] = df["hoggax_pct_sobre_total"] * 100.0

    return df


def style_header_row(ws: Worksheet, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = cast(Cell, ws.cell(row=row, column=col))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def heatmap(ws: Worksheet, rng: str):
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min",
            start_color="63BE7B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="F8696B",
        ),
    )


def safe_set(cell: Cell, value):
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def write_matrix_percent(
    ws: Worksheet,
    title: str,
    df_matrix_pct: pd.DataFrame,
    plazos: list[int],
    start_row: int,
    start_col: int,
):
    """
    df_matrix_pct: columnas: segmento + str(plazo) con valores en % (ej 60.0 para 60%)
    """
    r0 = start_row
    c0 = start_col

    tcell = cast(Cell, ws.cell(r0, c0))
    safe_set(tcell, title)
    tcell.font = Font(bold=True)
    tcell.fill = PatternFill("solid", fgColor="C6E0B4")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + len(plazos))

    # header
    ws.cell(r0 + 1, c0, value="segmento")
    for i, m in enumerate(plazos):
        ws.cell(r0 + 1, c0 + 1 + i, value=m)
    style_header_row(ws, r0 + 1, c0, c0 + len(plazos))

    # data
    for i, (_, row) in enumerate(df_matrix_pct.iterrows(), start=0):
        rr = r0 + 2 + i
        ws.cell(rr, c0, value=str(row["segmento"]))
        for k, m in enumerate(plazos):
            v = row.get(str(m), None)
            cell = cast(Cell, ws.cell(rr, c0 + 1 + k))
            fv = to_float(v)
            if fv is None:
                safe_set(cell, None)
            else:
                # Excel/Sheets: % requiere fracción
                safe_set(cell, fv / 100.0)
                cell.number_format = "0.00%"
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions[get_column_letter(c0)].width = 16
    for i in range(1, len(plazos) + 1):
        ws.column_dimensions[get_column_letter(c0 + i)].width = 10

    nrows = int(df_matrix_pct.shape[0])
    if nrows:
        top_left = ws.cell(r0 + 2, c0 + 1).coordinate
        bot_right = ws.cell(r0 + 1 + nrows, c0 + len(plazos)).coordinate
        heatmap(ws, f"{top_left}:{bot_right}")


def main():
    # ---------- INPUTS ----------
    finaer_jsonl = load_latest_jsonl("finaer_")

    hoggax_path = Path("data/hoggax_rates_long.csv")
    if not hoggax_path.exists():
        hoggax_path = Path("data/hoggax_rates.csv")  # fallback
    if not hoggax_path.exists():
        raise SystemExit("No encuentro data/hoggax_rates_long.csv ni data/hoggax_rates.csv")

    df_h_long = load_hoggax_rates_long(str(hoggax_path))

    # ---------- HOGGAX: descuento fijo 15% ----------
    df_h_long["hoggax_desc_pct"] = df_h_long["hoggax_desc_pct"].fillna(15.0)

    # ---------- FINAER (desde JSONL) ----------
    finaer_rows = []
    with open(finaer_jsonl, "r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rec = json.loads(line)

            s = rec.get("scenario") or {}
            alq = float(s.get("alquiler") or 0)
            exp = float(s.get("expensas") or 0)
            plazo_meses = int(s.get("meses") or 0)

            if plazo_meses not in PLAZOS_FINAER:
                continue

            total_base = alq + exp
            if total_base <= 0:
                continue

            planes = ((rec.get("normalized") or {}).get("planes")) or []
            p = pick_plan_contado(planes)

            monto_final = to_float(p.get("monto_final"))
            honorario = to_float(p.get("honorario_sin_descuentos"))
            desc_abs = to_float(p.get("descuento_aplicado"))
            fecha_desc = p.get("fecha_limite_descuento")

            if monto_final is None:
                continue

            # % sobre total (Alq+Exp) * plazo
            pct_sobre_total = (monto_final / (total_base * plazo_meses)) * 100.0

            # descuento % real
            desc_pct = None
            if honorario and honorario > 0 and desc_abs is not None:
                desc_pct = (desc_abs / honorario) * 100.0

            finaer_rows.append(
                {
                    "segmento": segmento(total_base),
                    "plazo_meses": plazo_meses,
                    "alquiler": alq,
                    "expensas": exp,
                    "total_base_$": total_base,
                    "finaer_precio_$": monto_final,
                    "finaer_pct_sobre_total": pct_sobre_total,
                    "finaer_honorario_sin_desc_$": honorario,
                    "finaer_desc_$": desc_abs,
                    "finaer_desc_pct": desc_pct,
                    "finaer_fecha_desc": fecha_desc,
                }
            )

    df_f = pd.DataFrame(finaer_rows)
    if df_f.empty:
        raise SystemExit("Finaer: no tengo filas (recordá: sólo 12/24/36). Corré la CLI con escenarios 12/24/36.")

    # ---------- DOMINIOS / VALORES QUE TRAE FINAER ----------
    def uniq_sorted(x: pd.Series):
        vals = [v for v in pd.unique(x.dropna())]
        try:
            return sorted([float(v) for v in vals])
        except Exception:
            return sorted([str(v) for v in vals])

    df_f_domains = (
        df_f.groupby(["segmento", "plazo_meses"], as_index=False)
        .agg(
            finaer_precio_values=("finaer_precio_$", uniq_sorted),
            finaer_pct_values=("finaer_pct_sobre_total", lambda s: [round(float(v), 6) for v in uniq_sorted(s)]),
            finaer_honorario_values=("finaer_honorario_sin_desc_$", uniq_sorted),
            finaer_desc_abs_values=("finaer_desc_$", uniq_sorted),
            finaer_desc_pct_values=("finaer_desc_pct", lambda s: [round(float(v), 6) for v in uniq_sorted(s)]),
            finaer_fecha_desc_values=("finaer_fecha_desc", lambda s: sorted([str(v) for v in pd.unique(s.dropna())])),
            n_escenarios=("finaer_precio_$", "count"),
        )
    )

    # ---------- PROMEDIOS FINAER (solo para comparativa/matrices) ----------
    df_f_avg = (
        df_f.groupby(["segmento", "plazo_meses"], as_index=False)
        .agg(
            **{
                "finaer_precio_$": ("finaer_precio_$", "mean"),
                "finaer_pct_sobre_total": ("finaer_pct_sobre_total", "mean"),
                "finaer_honorario_sin_desc_$": ("finaer_honorario_sin_desc_$", "mean"),
                "finaer_desc_$": ("finaer_desc_$", "mean"),
                "finaer_desc_pct": ("finaer_desc_pct", "mean"),
            }
        )
    )

    # ---------- HOGGAX (desde CSV manual) ----------
    df_h_avg = (
        df_h_long.groupby(["segmento", "plazo_meses"], as_index=False)
        .agg(
            **{
                "hoggax_precio_$": ("hoggax_precio_$", "mean"),
                "hoggax_pct_sobre_total": ("hoggax_pct_sobre_total", "mean"),
                "hoggax_desc_pct": ("hoggax_desc_pct", "mean"),
                "hoggax_cuota_$": ("hoggax_cuota_$", "mean"),
                "hoggax_cuotas": ("hoggax_cuotas", "mean"),
            }
        )
    )

    # ---------- MATRICES (% sobre total) ----------
    finaer_mat = (
        df_f_avg.pivot(index="segmento", columns="plazo_meses", values="finaer_pct_sobre_total")
        .reindex(index=SEGMENTOS, columns=PLAZOS_HOGGAX)
        .reset_index()
    )
    finaer_mat.columns = ["segmento"] + [str(m) for m in PLAZOS_HOGGAX]

    hoggax_mat = (
        df_h_avg.pivot(index="segmento", columns="plazo_meses", values="hoggax_pct_sobre_total")
        .reindex(index=SEGMENTOS, columns=PLAZOS_HOGGAX)
        .reset_index()
    )
    hoggax_mat.columns = ["segmento"] + [str(m) for m in PLAZOS_HOGGAX]

    # ---------- COMPARATIVA ----------
    comp = pd.merge(df_f_avg, df_h_avg, on=["segmento", "plazo_meses"], how="outer")

    comp["segmento"] = pd.Categorical(comp["segmento"], categories=SEGMENTOS, ordered=True)
    comp = comp.sort_values(["segmento", "plazo_meses"])

    # ---------- EXCEL ----------
    wb = Workbook()

    # Sheet 1: Matrices
    ws_m = cast(Worksheet, wb.active)
    ws_m.title = "Matrices"

    write_matrix_percent(
        ws_m,
        "Finaer (% sobre total garantía)",
        finaer_mat,
        PLAZOS_HOGGAX,
        start_row=1,
        start_col=1,
    )
    write_matrix_percent(
        ws_m,
        "Hoggax (% sobre total garantía)",
        hoggax_mat,
        PLAZOS_HOGGAX,
        start_row=1,
        start_col=8,
    )
    ws_m.freeze_panes = "A3"

    # Sheet 2: Comparativa (enfocada en monto_final y honorario_sin_descuentos)
    ws_c = cast(Worksheet, wb.create_sheet("Comparativa"))

    headers = [
        "segmento",
        "plazo_meses",
        "finaer_monto_final_$",
        "finaer_honorario_sin_desc_$",
        "finaer_desc_%",
        "hoggax_desc_%",
        "dif_desc_puntos_(F-H)",
    ]
    ws_c.append(headers)
    style_header_row(ws_c, 1, 1, len(headers))
    ws_c.freeze_panes = "A2"
    ws_c.auto_filter.ref = ws_c.dimensions

    widths = {
        "A": 16,
        "B": 11,
        "C": 18,
        "D": 22,
        "E": 14,
        "F": 14,
        "G": 18,
    }
    for col, w in widths.items():
        ws_c.column_dimensions[col].width = w

    for _, r in comp.iterrows():
        seg = str(r["segmento"]) if pd.notna(r["segmento"]) else ""
        plazo = int(r["plazo_meses"]) if pd.notna(r["plazo_meses"]) else None

        f_price = to_float(r.get("finaer_precio_$"))
        f_honorario = to_float(r.get("finaer_honorario_sin_desc_$"))
        f_desc_pct = to_float(r.get("finaer_desc_pct"))

        h_desc_pct = to_float(r.get("hoggax_desc_pct"))

        ws_c.append(
            [
                seg,
                plazo,
                f_price,
                f_honorario,
                f_desc_pct,
                h_desc_pct,
                None,
            ]
        )

    for rr in range(2, ws_c.max_row + 1):
        # moneda: C (monto_final), D (honorario_sin_descuentos)
        for col in (3, 4):
            cell = cast(Cell, ws_c.cell(rr, col))
            if isinstance(cell, MergedCell):
                continue
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="center")

        # %: E (finaer_desc), F (hoggax_desc), G (dif_desc) - guardamos como fracción
        for col in (5, 6, 7):
            cell = cast(Cell, ws_c.cell(rr, col))
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.value = float(v) / 100.0
            cell.number_format = "0.00%"
            cell.alignment = Alignment(horizontal="center")

        # diferencia en puntos porcentuales entre Finaer y Hoggax
        cast(Cell, ws_c.cell(rr, 7)).value = f'=IF(OR(E{rr}="",F{rr}=""),"",E{rr}-F{rr})'

    heatmap(ws_c, f"G2:G{ws_c.max_row}")

    # Sheet 3: Finaer_Valores (dominios reales)
    ws_v = cast(Worksheet, wb.create_sheet("Finaer_Valores"))
    headers_v = [
        "segmento",
        "plazo_meses",
        "n_escenarios",
        "finaer_desc_pct_values",
        "finaer_desc_abs_values",
        "finaer_honorario_values",
        "finaer_precio_values",
        "finaer_pct_values",
        "finaer_fecha_desc_values",
    ]
    ws_v.append(headers_v)
    style_header_row(ws_v, 1, 1, len(headers_v))
    ws_v.freeze_panes = "A2"
    ws_v.auto_filter.ref = ws_v.dimensions

    for i, w in enumerate([16, 11, 12, 28, 28, 28, 28, 28, 28], start=1):
        ws_v.column_dimensions[get_column_letter(i)].width = w

    df_f_domains = df_f_domains.sort_values(["segmento", "plazo_meses"])
    for _, r in df_f_domains.iterrows():
        ws_v.append(
            [
                str(r["segmento"]),
                int(r["plazo_meses"]),
                int(r["n_escenarios"]),
                str(r["finaer_desc_pct_values"]),
                str(r["finaer_desc_abs_values"]),
                str(r["finaer_honorario_values"]),
                str(r["finaer_precio_values"]),
                str(r["finaer_pct_values"]),
                str(r["finaer_fecha_desc_values"]),
            ]
        )

    out = Path("output") / f"compare_exact_finaer_{finaer_jsonl.stem}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote Excel -> {out}")


if __name__ == "__main__":
    main()
