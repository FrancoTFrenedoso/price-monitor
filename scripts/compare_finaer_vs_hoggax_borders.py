# scripts/compare_finaer_vs_hoggax_borders.py
from __future__ import annotations

from pathlib import Path
from typing import Optional, cast

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------- Config ----------------
# Si querés fijar el archivo de Finaer, dejalo así.
# Si lo ponés None, toma el último output/finaer_*.xlsx (crudo) automáticamente.
FINAER_XLSX_EXACT: Optional[Path] = None
# Ejemplo para usar el crudo:
# FINAER_XLSX_EXACT = Path("output/finaer_2026-02-11T135228Z.xlsx")

# Hoggax generado por API (scripts/fetch_hoggax_quotes.py)
HOGGAX_API_LONG = Path("output/hoggax_rates_long.csv")

OUT = Path("output/compare_borders_finaer_vs_hoggax.xlsx")

TARGET_ALQ_EXP = {499_999, 799_999, 801_000}
TARGET_MESES = {24, 36}
TARGET_CUOTAS = {1, 3}

# Finaer: 20% off SOLO contado (1 pago)
FINAER_CONTADO_DESC_PCT = 20.0


# ---------------- Helpers ----------------
def seg_label(alq_exp: float) -> str:
    if alq_exp <= 500_000:
        return "hasta_500k"
    if alq_exp <= 800_000:
        return "500_800k"
    return "mayor_800k"


def parse_num(x) -> Optional[float]:
    """
    Convierte números que pueden venir como:
    - 899998.2
    - "899.998,2"
    - "$899.998"
    - "899,998.2" (raro)
    """
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(" ", "")
    # si tiene coma y punto, asumimos formato AR: 1.234.567,89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # si solo tiene coma, puede ser decimal
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        # si solo tiene puntos, pueden ser miles: "1.234.567"
        # en ese caso dejamos puntos si parece decimal (solo un punto con 1-2-3 decimales),
        # pero si hay más de uno, son miles.
        if s.count(".") > 1:
            s = s.replace(".", "")
    try:
        return float(s)
    except Exception:
        return None


def pick_latest_finaer_xlsx() -> Path:
    out_dir = Path("output")
    cands = []
    for p in out_dir.glob("finaer_*.xlsx"):
        name = p.name.lower()
        if name.startswith("finaer_") and "compare" not in name and "summary" not in name and "matrices" not in name:
            cands.append(p)
    if not cands:
        raise SystemExit("No encontré output/finaer_*.xlsx. Corré primero el CLI que genera el Excel de Finaer.")
    return max(cands, key=lambda p: p.stat().st_mtime)


def style_header_row(ws: Worksheet, row: int, start_col: int, end_col: int):
    fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(start_col, end_col + 1):
        cell = cast(Cell, ws.cell(row=row, column=col))
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def money(cell: Cell):
    cell.number_format = '"$"#,##0'
    cell.alignment = Alignment(horizontal="center")


def pct(cell: Cell):
    cell.number_format = "0.00%"
    cell.alignment = Alignment(horizontal="center")


def heatmap(ws: Worksheet, rng: str):
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min",
            start_color="63BE7B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="F8696B",
        ),
    )


def main():
    finaer_xlsx = FINAER_XLSX_EXACT or pick_latest_finaer_xlsx()
    if not finaer_xlsx.exists():
        raise SystemExit(f"No existe {finaer_xlsx}")

    if not HOGGAX_API_LONG.exists():
        raise SystemExit(f"No existe {HOGGAX_API_LONG}. Corré: python scripts/fetch_hoggax_quotes.py")

    # ---------------- FINaer (acepta crudo o limpio) ----------------
    df = pd.read_excel(finaer_xlsx)
    df.columns = [str(c).strip() for c in df.columns]

    # columnas base
    required_base = {"alq_exp", "meses", "cuotas"}
    missing_base = required_base - set(df.columns)
    if missing_base:
        raise SystemExit(f"Faltan columnas base en {finaer_xlsx.name}: {sorted(missing_base)}")

    # normalizar base
    for c in ["alq_exp", "meses", "cuotas"]:
        df[c] = df[c].apply(parse_num)

    df["alq_exp"] = df["alq_exp"].astype("Int64")
    df["meses"] = df["meses"].astype("Int64")
    df["cuotas"] = df["cuotas"].astype("Int64")

    # lista: crudo o limpio
    if "honorario_sin_desc" in df.columns:
        df["finaer_lista"] = df["honorario_sin_desc"].apply(parse_num)
    elif "honorario_sin_descuentos" in df.columns:
        df["finaer_lista"] = df["honorario_sin_descuentos"].apply(parse_num)
    elif "finaer_lista" in df.columns:
        df["finaer_lista"] = df["finaer_lista"].apply(parse_num)
    else:
        raise SystemExit(
            f"Falta columna de lista en {finaer_xlsx.name}: "
            f"'honorario_sin_desc'/'honorario_sin_descuentos' o 'finaer_lista'"
        )

    # opcionales del crudo
    for c in ["total_final", "monto_cuota", "anticipo", "finaer_total_transfer", "finaer_cuota_equiv"]:
        if c in df.columns:
            df[c] = df[c].apply(parse_num)

    # filtros bordes
    df = df[df["alq_exp"].isin(TARGET_ALQ_EXP)]
    df = df[df["meses"].isin(TARGET_MESES)]
    df = df[df["cuotas"].isin(TARGET_CUOTAS)].copy()

    # segmento
    if "segmento" not in df.columns:
        df["segmento"] = df["alq_exp"].astype(float).apply(seg_label)

    # --- Finaer: regla correcta ---
    # contado: 20% off (total = lista * 0.80)
    # 3 cuotas: sin descuento (total = lista)
    df["finaer_transfer_desc_pct"] = 0.0
    df.loc[df["cuotas"] == 1, "finaer_transfer_desc_pct"] = FINAER_CONTADO_DESC_PCT

    # si ya venía calculado, lo pisamos igual para evitar basura por rounding en previos
    df["finaer_total_transfer"] = df["finaer_lista"]
    df.loc[df["cuotas"] == 1, "finaer_total_transfer"] = (
        df.loc[df["cuotas"] == 1, "finaer_lista"] * (1.0 - FINAER_CONTADO_DESC_PCT / 100.0)
    )

    df["finaer_cuota_equiv"] = df["finaer_total_transfer"] / df["cuotas"].astype(float)

    # ---------------- Hoggax (API long por input exacto + plan) ----------------
    dh = pd.read_csv(HOGGAX_API_LONG)
    dh.columns = [str(c).strip() for c in dh.columns]

    needed_h = {"alq_exp", "meses", "cuotas", "hoggax_sin_desc", "hoggax_total_web"}
    missing_h = needed_h - set(dh.columns)
    if missing_h:
        raise SystemExit(f"CSV Hoggax inválido. Faltan columnas: {sorted(missing_h)}")

    for c in ["alq_exp", "meses", "cuotas", "hoggax_sin_desc", "hoggax_total_web", "hoggax_monto_cuota"]:
        if c in dh.columns:
            dh[c] = dh[c].apply(parse_num)

    dh["alq_exp"] = dh["alq_exp"].astype("Int64")
    dh["meses"] = dh["meses"].astype("Int64")
    dh["cuotas"] = dh["cuotas"].astype("Int64")

    df = df.merge(dh, on=["alq_exp", "meses", "cuotas"], how="left")

    df["hoggax_lista"] = df["hoggax_sin_desc"]
    df["hoggax_total_transfer"] = df["hoggax_total_web"]
    df["hoggax_cuota_equiv"] = df["hoggax_total_transfer"] / df["cuotas"].astype(float)

    # % descuento efectivo Hoggax (informativo, en %)
    df["hoggax_transfer_desc_pct"] = 0.0
    mask = (
        (df["cuotas"] == 1)
        & df["hoggax_lista"].notna()
        & df["hoggax_total_transfer"].notna()
        & (df["hoggax_lista"] > 0)
    )
    df.loc[mask, "hoggax_transfer_desc_pct"] = (
        (1.0 - (df.loc[mask, "hoggax_total_transfer"] / df.loc[mask, "hoggax_lista"])) * 100.0
    )

    # difs
    df["dif_lista_$"] = df["finaer_lista"] - df["hoggax_lista"]
    df["dif_total_transfer_$"] = df["finaer_total_transfer"] - df["hoggax_total_transfer"]

    df = df.sort_values(["alq_exp", "meses", "cuotas"]).reset_index(drop=True)

    # columnas finales (agrega extras si existen)
    out_cols = [
        "segmento",
        "alq_exp",
        "meses",
        "cuotas",
        "finaer_lista",
        "finaer_transfer_desc_pct",
        "finaer_total_transfer",
        "finaer_cuota_equiv",
    ]
    for c in ["total_final", "monto_cuota", "anticipo"]:
        if c in df.columns:
            out_cols.append(c)

    out_cols += [
        "hoggax_lista",
        "hoggax_transfer_desc_pct",
        "hoggax_total_transfer",
        "hoggax_cuota_equiv",
        "dif_lista_$",
        "dif_total_transfer_$",
    ]

    df_out = df[out_cols].copy()

    # ---------------- Excel formatting ----------------
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Comparativa"
    ws.freeze_panes = "A2"

    ws.append(out_cols)
    style_header_row(ws, 1, 1, len(out_cols))
    ws.auto_filter.ref = ws.dimensions

    # anchos (más grandes para columnas pedidas)
    wide_cols = {
        "finaer_transfer_desc_pct": 26,
        "monto_cuota": 20,
        "anticipo": 20,
        "hoggax_cuota_equiv": 22,
        "cuotas": 12,
    }

    for i, col in enumerate(out_cols, start=1):
        letter = get_column_letter(i)
        if col in wide_cols:
            ws.column_dimensions[letter].width = wide_cols[col]
        elif col == "segmento":
            ws.column_dimensions[letter].width = 16
        elif col in ("alq_exp",):
            ws.column_dimensions[letter].width = 14
        elif col in ("meses",):
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 18

    # cargar filas
    for _, r in df_out.iterrows():
        ws.append([r[c] if c == "segmento" else parse_num(r[c]) for c in out_cols])

    col_idx = {name: i + 1 for i, name in enumerate(out_cols)}

    money_cols = set(
        [
            "alq_exp",
            "finaer_lista",
            "finaer_total_transfer",
            "finaer_cuota_equiv",
            "hoggax_lista",
            "hoggax_total_transfer",
            "hoggax_cuota_equiv",
            "dif_lista_$",
            "dif_total_transfer_$",
            "total_final",
            "monto_cuota",
            "anticipo",
        ]
    )
    pct_cols = {"finaer_transfer_desc_pct", "hoggax_transfer_desc_pct"}

    # bandas por segmento
    fill_a = PatternFill("solid", fgColor="FFFFFF")
    fill_b = PatternFill("solid", fgColor="F2F2F2")

    for row in range(2, ws.max_row + 1):
        if "meses" in col_idx:
            ws.cell(row, col_idx["meses"]).alignment = Alignment(horizontal="center")
        if "cuotas" in col_idx:
            ws.cell(row, col_idx["cuotas"]).alignment = Alignment(horizontal="center")

        for c in out_cols:
            if c in money_cols:
                cell = cast(Cell, ws.cell(row, col_idx[c]))
                if isinstance(cell, MergedCell) or cell.coordinate in ws.merged_cells:
                    continue
                money(cell)

            if c in pct_cols:
                cell = cast(Cell, ws.cell(row, col_idx[c]))
                v = cell.value
                if isinstance(v, (int, float)) and v is not None:
                    cell.value = float(v) / 100.0
                pct(cell)

        seg = ws.cell(row, col_idx["segmento"]).value
        fill = fill_b if seg in ("hasta_500k", "mayor_800k") else fill_a
        for c in range(1, len(out_cols) + 1):
            ws.cell(row, c).fill = fill

    # heatmap sobre dif_total_transfer_$
    if "dif_total_transfer_$" in col_idx:
        start = ws.cell(2, col_idx["dif_total_transfer_$"]).coordinate
        end = ws.cell(ws.max_row, col_idx["dif_total_transfer_$"]).coordinate
        heatmap(ws, f"{start}:{end}")

    wb.save(OUT)
    print("Read Finaer ->", finaer_xlsx)
    print("Read Hoggax ->", HOGGAX_API_LONG)
    print("Wrote Excel ->", OUT)


if __name__ == "__main__":
    main()
