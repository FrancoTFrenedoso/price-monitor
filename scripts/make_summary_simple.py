from __future__ import annotations

import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill


def seg(a: float) -> str:
    if a <= 500_000:
        return "hasta 500k"
    if a <= 800_000:
        return "500k-800k"
    return "mayor_800k"


def load_latest():
    out = Path("output")
    files = sorted(out.glob("*.jsonl"))
    if not files:
        raise SystemExit("No hay jsonl en output/")
    f = files[-1]

    rows = []
    for line in open(f, encoding="utf-8"):
        if not line.strip():
            continue
        r = json.loads(line)
        s = r["scenario"]
        for p in r["normalized"]["planes"]:
            rows.append({
                "scenario_id": r["scenario_id"],
                "alquiler": s["alquiler"],
                "meses": s["meses"],
                "cuotas": p["cuotas"],
                "monto_final": p["monto_final"],
                "pct_desc": p.get("pct_descuento_real"),
                "pct_contrato": p.get("pct_sobre_total_alq_exp"),
                "costo_mensual": p.get("costo_mensual_equiv"),
            })
    return pd.DataFrame(rows), f.stem


def pick_plan(df):
    df = df.sort_values(["scenario_id", "cuotas"])
    return df.groupby("scenario_id", as_index=False).first()


def apply_formatting(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # escala de color para precio promedio
    price_rule = ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B"
    )
    ws.conditional_formatting.add("C2:C200", price_rule)

    # escala para % descuento
    desc_rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"
    )
    ws.conditional_formatting.add("D2:D200", desc_rule)

    # pintar segmentos
    fills = {
        "hasta 500k": PatternFill("solid", fgColor="D9E1F2"),
        "500k-800k": PatternFill("solid", fgColor="E2EFDA"),
        "mayor_800k": PatternFill("solid", fgColor="FCE4D6"),
    }

    for row in ws.iter_rows(min_row=2, max_col=1):
        val = row[0].value
        if val in fills:
            row[0].fill = fills[val]

    # autofilter y freeze
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)


def main():
    df, stem = load_latest()
    df = pick_plan(df)
    df["segmento"] = df["alquiler"].apply(seg)

    g = (
        df.groupby(["segmento", "meses"], as_index=False)
        .agg(
            precio_prom=("monto_final", "mean"),
            descuento_prom=("pct_desc", "mean"),
            pct_contrato_prom=("pct_contrato", "mean"),
            costo_mensual_prom=("costo_mensual", "mean"),
            n=("scenario_id", "nunique"),
        )
        .sort_values(["segmento", "meses"])
    )

    out = Path("output") / f"summary_simple_{stem}.xlsx"
    g.to_excel(out, index=False)

    apply_formatting(out)
    print("Wrote", out)


if __name__ == "__main__":
    main()
